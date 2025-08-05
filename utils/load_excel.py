import openpyxl

def extract_text_from_excel(path):
    wb = openpyxl.load_workbook(path)
    text = ''
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_text = [str(cell) if cell else '' for cell in row]
            text += ' | '.join(row_text) + '\n'
    return text
